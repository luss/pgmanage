'''
The MIT License (MIT)

Copyright (c) 2014-2019 William Ivanski

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
'''

import os
import csv
import openpyxl
from collections import OrderedDict
import tempfile

import app.include.Spartacus as Spartacus

class Exception(Exception):
    pass


class DataFileWriter(object):
    def __init__(self, p_filename, p_fieldnames=None, p_encoding='utf-8', p_delimiter=';', p_lineterminator='\n', skip_headers=False):
        v_tmp = p_filename.split('.')
        if len(v_tmp) > 1:
            self.v_extension = v_tmp[-1].lower()
        else:
            self.v_extension = 'csv'
        if self.v_extension == 'txt' or self.v_extension == 'out':
            self.v_extension = 'csv'
        self.v_filename = p_filename
        self.v_file = None
        self.v_header = p_fieldnames # Can't be empty for CSV
        self.v_encoding = p_encoding
        self.v_delimiter = p_delimiter
        self.v_lineterminator = p_lineterminator
        self.v_currentrow = 1
        self.v_open = False
        self.skip_headers = skip_headers
    def Open(self):
        try:
            if self.v_extension == 'csv':
                self.v_file = open(self.v_filename, 'w', encoding=self.v_encoding)
                self.v_object = csv.writer(self.v_file, delimiter=self.v_delimiter, lineterminator=self.v_lineterminator)
                if not self.skip_headers:
                    self.v_object.writerow(self.v_header)
                self.v_open = True
            elif self.v_extension == 'xlsx':
                self.v_object = openpyxl.Workbook(write_only=True)
                self.v_open = True
            else:
                raise Spartacus.Utils.Exception('File extension "{0}" not supported.'.format(self.v_extension))
        except Spartacus.Utils.Exception as exc:
            raise exc
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))
    def Write(self, p_datatable, p_sheetname=None):
        try:
            if not self.v_open:
                raise Spartacus.Utils.Exception('You need to call Open() first.')
            if self.v_extension == 'csv':
                for v_row in p_datatable.Rows:
                    self.v_object.writerow(v_row)
            else:
                if self.v_currentrow == 1:
                    if p_sheetname:
                        v_worksheet = self.v_object.create_sheet(p_sheetname)
                    else:
                        v_worksheet = self.v_object.create_sheet()
                    if not self.skip_headers:
                        v_worksheet.append(p_datatable.Columns)
                        self.v_currentrow = self.v_currentrow + 1
                else:
                    v_worksheet = self.v_object.active
                for r in range(0, len(p_datatable.Rows)):
                    v_row = []
                    for c in range(0, len(p_datatable.Columns)):
                        v_row.append(p_datatable.Rows[r][c])
                    v_worksheet.append(v_row)
                self.v_currentrow = self.v_currentrow + len(p_datatable.Rows)
        except Spartacus.Utils.Exception as exc:
            raise exc
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))
    def Flush(self):
        try:
            if not self.v_open:
                raise Spartacus.Utils.Exception('You need to call Open() first.')
            if self.v_extension == 'csv':
                self.v_file.close()
            else:
                self.v_object.save(self.v_filename)
        except Spartacus.Utils.Exception as exc:
            raise exc
        except Exception as exc:
            raise Spartacus.Utils.Exception(str(exc))
